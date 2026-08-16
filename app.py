import io
import csv
import openpyxl
from openpyxl import Workbook
from flask import (
    Flask,
    make_response,
    render_template,
    request,
    send_file,
    redirect,
    url_for,
    Response,
)
from modules.db import get_connection

app = Flask(__name__)


@app.route("/")
def home():

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT COUNT(*) FROM customers")
    customers_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM materials")
    materials_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM sales_orders")
    sales_orders_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM support_requests")
    support_requests_count = cursor.fetchone()[0]

    cursor.close()
    connection.close()

    return render_template(
        "index.html",
        customers_count=customers_count,
        materials_count=materials_count,
        sales_orders_count=sales_orders_count,
        support_requests_count=support_requests_count,
    )


@app.route("/customers")
def customers():

    search = request.args.get("search", "").strip()
    country = request.args.get("country", "").strip()
    status = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "customer_id")
    sort_order = request.args.get("sort_order", "asc")

    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    # Allowed page sizes
    allowed_page_sizes = [10, 25, 50, 100]

    if per_page not in allowed_page_sizes:
        per_page = 10

    if page < 1:
        page = 1

    # Allowed columns for sorting
    sort_options = {
        "customer_id": "customer_id",
        "customer_name": "CAST(REGEXP_REPLACE(customer_name, '[^0-9]', '', 'g') AS INTEGER)",
        "country": "country",
        "status": "status",
    }

    # Prevent invalid column names
    order_column = sort_options.get(sort_by, "customer_id")

    # Only allow ASC / DESC
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    # Main query
    query = """
        SELECT customer_id,
               customer_name,
               country,
               status
        FROM customers
        WHERE 1=1
    """

    params = []

    # Search
    if search:
        query += """
            AND (
                customer_id ILIKE %s
                OR customer_name ILIKE %s
                OR country ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])

    # Country filter
    if country:
        query += " AND country = %s"
        params.append(country)

    # Status filter
    if status:
        query += " AND status = %s"
        params.append(status)

    # Count filtered records
    count_query = """
        SELECT COUNT(*)
        FROM customers
        WHERE 1=1
    """

    count_params = []

    # Search for count
    if search:
        count_query += """
            AND (
                customer_id ILIKE %s
                OR customer_name ILIKE %s
                OR country ILIKE %s
                OR status ILIKE %s
            )
        """

        count_params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # Country filter for count
    if country:
        count_query += " AND country = %s"
        count_params.append(country)

    # Status filter for count
    if status:
        count_query += " AND status = %s"
        count_params.append(status)

    cursor.execute(count_query, count_params)

    total_records = cursor.fetchone()[0]

    # Calculate total pages
    total_pages = max(1, (total_records + per_page - 1) // per_page)

    # Prevent invalid page number
    if page > total_pages:
        page = total_pages

    # Calculate offset
    offset = (page - 1) * per_page

    # Sorting + Pagination
    query += f"""
        ORDER BY {order_column} {order_direction}
        LIMIT %s OFFSET %s
    """

    params.extend([per_page, offset])

    cursor.execute(query, params)

    customers = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "customers.html",
        customers=customers,
        search=search,
        country=country,
        status=status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
        total_records=total_records,
        total_pages=total_pages,
        current_endpoint="customers",
    )


@app.route("/customers/create", methods=["GET", "POST"])
def create_customer():

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    country_filter = request.args.get("country", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "customer_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    per_page = request.args.get("per_page", "10").strip()
    page = request.args.get("page", "1").strip()

    # -----------------------------------
    # POST - Create customer
    # -----------------------------------

    if request.method == "POST":

        customer_id = request.form.get("customer_id", "").strip()
        customer_name = request.form.get("customer_name", "").strip()
        country = request.form.get("country", "").strip()
        status = request.form.get("status", "").strip()

        # -----------------------------------
        # Basic validation
        # -----------------------------------

        if not customer_id or not customer_name or not country or not status:

            return render_template(
                "customer_form.html",
                error="All fields are required.",
                customer_id=customer_id,
                customer_name=customer_name,
                country=country,
                status=status,
                search=search,
                country_filter=country_filter,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                per_page=per_page,
                page=page,
            )

        connection = get_connection()
        cursor = connection.cursor()

        # -----------------------------------
        # Check duplicate Customer ID
        # -----------------------------------

        cursor.execute(
            """
            SELECT customer_id
            FROM customers
            WHERE customer_id = %s
            """,
            (customer_id,),
        )

        existing_customer = cursor.fetchone()

        if existing_customer:

            cursor.close()
            connection.close()

            return render_template(
                "customer_form.html",
                error="Customer ID already exists.",
                customer_id=customer_id,
                customer_name=customer_name,
                country=country,
                status=status,
                search=search,
                country_filter=country_filter,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                per_page=per_page,
                page=page,
            )

        # -----------------------------------
        # Insert customer
        # -----------------------------------

        cursor.execute(
            """
            INSERT INTO customers
                (customer_id, customer_name, country, status)
            VALUES
                (%s, %s, %s, %s)
            """,
            (customer_id, customer_name, country, status),
        )

        connection.commit()

        cursor.close()
        connection.close()

        # -----------------------------------
        # Return to same filtered/sorted page
        # -----------------------------------

        return redirect(
            url_for(
                "customers",
                search=search,
                country=country_filter,
                status=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                per_page=per_page,
                page=page,
            )
        )

    # -----------------------------------
    # GET - Show create form
    # -----------------------------------

    return render_template(
        "customer_form.html",
        search=search,
        country_filter=country_filter,
        status_filter=status_filter,
        sort_by=sort_by,
        sort_order=sort_order,
        per_page=per_page,
        page=page,
    )


@app.route("/customers/<customer_id>/edit", methods=["GET", "POST"])
def edit_customer(customer_id):

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    country = request.args.get("country", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "customer_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    per_page = request.args.get("per_page", "10").strip()
    page = request.args.get("page", "1").strip()

    # -----------------------------------
    # GET existing customer
    # -----------------------------------

    if request.method == "GET":

        cursor.execute(
            """
            SELECT customer_id,
                   customer_name,
                   country,
                   status
            FROM customers
            WHERE customer_id = %s
        """,
            (customer_id,),
        )

        customer = cursor.fetchone()

        cursor.close()
        connection.close()

        if not customer:
            return "Customer not found", 404

        return render_template(
            "edit_customer.html",
            customer=customer,
            search=search,
            country_filter=country,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            per_page=per_page,
            page=page,
        )

    # -----------------------------------
    # POST - Update customer
    # -----------------------------------

    customer_name = request.form.get("customer_name", "").strip()
    new_country = request.form.get("country", "").strip()
    new_status = request.form.get("status", "").strip()

    # Validation
    if not customer_name or not new_country or not new_status:

        cursor.execute(
            """
            SELECT customer_id,
                   customer_name,
                   country,
                   status
            FROM customers
            WHERE customer_id = %s
        """,
            (customer_id,),
        )

        customer = cursor.fetchone()

        cursor.close()
        connection.close()

        return render_template(
            "edit_customer.html",
            customer=customer,
            error="All fields are required.",
            search=search,
            country_filter=country,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            per_page=per_page,
            page=page,
        )

    # -----------------------------------
    # Update database
    # -----------------------------------

    cursor.execute(
        """
        UPDATE customers
        SET customer_name = %s,
            country = %s,
            status = %s
        WHERE customer_id = %s
    """,
        (customer_name, new_country, new_status, customer_id),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "customers",
            search=search,
            country=country,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            per_page=per_page,
            page=page,
        )
    )


@app.route("/customers/<customer_id>/delete", methods=["POST"])
def delete_customer(customer_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    country = request.args.get("country", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "customer_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    per_page = request.args.get("per_page", "10").strip()
    page = request.args.get("page", "1").strip()

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        DELETE FROM customers
        WHERE customer_id = %s
    """,
        (customer_id,),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "customers",
            search=search,
            country=country,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            per_page=per_page,
            page=page,
        )
    )


@app.route("/customers/export/csv")
def export_customers_csv():

    search = request.args.get("search", "").strip()
    country = request.args.get("country", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "customer_id")
    sort_order = request.args.get("sort_order", "asc")

    # Allowed sorting columns
    sort_options = {
        "customer_id": "customer_id",
        "customer_name": "CAST(REGEXP_REPLACE(customer_name, '[^0-9]', '', 'g') AS INTEGER)",
        "country": "country",
        "status": "status",
    }

    order_column = sort_options.get(sort_by, "customer_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT customer_id,
               customer_name,
               country,
               status
        FROM customers
        WHERE 1=1
    """

    params = []

    # Search
    if search:
        query += """
            AND (
                customer_id ILIKE %s
                OR customer_name ILIKE %s
                OR country ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])

    # Country filter
    if country:
        query += " AND country = %s"
        params.append(country)

    # Status filter
    if status:
        query += " AND status = %s"
        params.append(status)

    # Sorting
    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Customer ID", "Customer Name", "Country", "Status"])

    writer.writerows(rows)

    response = make_response(output.getvalue())

    response.headers["Content-Disposition"] = "attachment; filename=customers.csv"

    response.headers["Content-Type"] = "text/csv"

    return response


@app.route("/customers/export/excel")
def export_customers_excel():

    search = request.args.get("search", "").strip()
    country = request.args.get("country", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "customer_id")
    sort_order = request.args.get("sort_order", "asc")

    # Allowed sorting columns
    sort_options = {
        "customer_id": "customer_id",
        "customer_name": "CAST(REGEXP_REPLACE(customer_name, '[^0-9]', '', 'g') AS INTEGER)",
        "country": "country",
        "status": "status",
    }

    order_column = sort_options.get(sort_by, "customer_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT customer_id,
               customer_name,
               country,
               status
        FROM customers
        WHERE 1=1
    """

    params = []

    # Search
    if search:
        query += """
            AND (
                customer_id ILIKE %s
                OR customer_name ILIKE %s
                OR country ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])

    # Country filter
    if country:
        query += " AND country = %s"
        params.append(country)

    # Status filter
    if status:
        query += " AND status = %s"
        params.append(status)

    # Sorting
    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Customers"

    # Header
    worksheet.append(["Customer ID", "Customer Name", "Country", "Status"])

    # Data
    for row in rows:
        worksheet.append(row)

    # Auto-size columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # Save workbook to memory
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="customers.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/materials")
def materials():

    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    plant = request.args.get("plant", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "material_id")
    sort_order = request.args.get("sort_order", "asc")

    # Pagination
    page = request.args.get("page", 1, type=int)

    # Rows per page
    per_page = request.args.get("per_page", 10, type=int)

    # Allow only valid page sizes
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    if page < 1:
        page = 1

    # Allowed columns for sorting
    sort_options = {
        "material_id": "material_id",
        "material_name": "CAST(REGEXP_REPLACE(material_name, '[^0-9]', '', 'g') AS INTEGER)",
        "category": "category",
        "plant": "plant",
        "price": "price",
        "status": "status",
    }

    # Prevent invalid column names
    order_column = sort_options.get(sort_by, "material_id")

    # Only allow ASC / DESC
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Total records count
    # -----------------------------------

    count_query = """
        SELECT COUNT(*)
        FROM materials
        WHERE 1=1
    """

    count_params = []

    # Search
    if search:
        count_query += """
            AND (
                material_id ILIKE %s
                OR material_name ILIKE %s
                OR category ILIKE %s
                OR plant ILIKE %s
                OR status ILIKE %s
            )
        """

        count_params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # Category filter
    if category:
        count_query += " AND category = %s"
        count_params.append(category)

    # Plant filter
    if plant:
        count_query += " AND plant = %s"
        count_params.append(plant)

    # Status filter
    if status:
        count_query += " AND status = %s"
        count_params.append(status)

    cursor.execute(count_query, count_params)

    total_records = cursor.fetchone()[0]

    # -----------------------------------
    # Calculate total pages
    # -----------------------------------

    total_pages = max(1, (total_records + per_page - 1) // per_page)

    # Prevent invalid page number
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    # -----------------------------------
    # Main query
    # -----------------------------------

    query = """
        SELECT material_id,
               material_name,
               category,
               plant,
               price,
               status
        FROM materials
        WHERE 1=1
    """

    params = []

    # Search
    if search:
        query += """
            AND (
                material_id ILIKE %s
                OR material_name ILIKE %s
                OR category ILIKE %s
                OR plant ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # Category filter
    if category:
        query += " AND category = %s"
        params.append(category)

    # Plant filter
    if plant:
        query += " AND plant = %s"
        params.append(plant)

    # Status filter
    if status:
        query += " AND status = %s"
        params.append(status)

    # Sorting + Pagination
    query += f"""
        ORDER BY {order_column} {order_direction}
        LIMIT %s OFFSET %s
    """

    params.extend([per_page, offset])

    cursor.execute(query, params)

    materials = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "materials.html",
        materials=materials,
        search=search,
        category=category,
        plant=plant,
        status=status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
        total_records=total_records,
        total_pages=total_pages,
    )


@app.route("/materials/create", methods=["GET", "POST"])
def create_material():

    # Preserve current Materials page state
    search = request.args.get("search", "").strip()
    category_filter = request.args.get("category", "").strip()
    plant_filter = request.args.get("plant", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "material_id")
    sort_order = request.args.get("sort_order", "asc")
    page = request.args.get("page", "1")
    per_page = request.args.get("per_page", "10")

    if request.method == "POST":

        material_id = request.form.get("material_id", "").strip()
        material_name = request.form.get("material_name", "").strip()
        category = request.form.get("category", "").strip()
        plant = request.form.get("plant", "").strip()
        price = request.form.get("price", "").strip()
        status = request.form.get("status", "").strip()

        # Preserve page state from hidden fields
        search = request.form.get("search", search)
        category_filter = request.form.get("category_filter", category_filter)
        plant_filter = request.form.get("plant_filter", plant_filter)
        status_filter = request.form.get("status_filter", status_filter)
        sort_by = request.form.get("sort_by", sort_by)
        sort_order = request.form.get("sort_order", sort_order)
        page = request.form.get("page", page)
        per_page = request.form.get("per_page", per_page)

        # Basic validation
        if (
            not material_id
            or not material_name
            or not category
            or not plant
            or not price
            or not status
        ):
            return render_template(
                "material_form.html",
                error="All fields are required.",
                material_id=material_id,
                material_name=material_name,
                category=category,
                plant=plant,
                price=price,
                status=status,
                search=search,
                category_filter=category_filter,
                plant_filter=plant_filter,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )

        # Price validation
        try:
            price = float(price)
        except ValueError:
            return render_template(
                "material_form.html",
                error="Price must be a valid number.",
                material_id=material_id,
                material_name=material_name,
                category=category,
                plant=plant,
                price=price,
                status=status,
                search=search,
                category_filter=category_filter,
                plant_filter=plant_filter,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )

        connection = get_connection()
        cursor = connection.cursor()

        # Check duplicate Material ID
        cursor.execute(
            "SELECT material_id FROM materials WHERE material_id = %s", (material_id,)
        )

        existing_material = cursor.fetchone()

        if existing_material:
            cursor.close()
            connection.close()

            return render_template(
                "material_form.html",
                error="Material ID already exists.",
                material_id=material_id,
                material_name=material_name,
                category=category,
                plant=plant,
                price=price,
                status=status,
                search=search,
                category_filter=category_filter,
                plant_filter=plant_filter,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )

        # Insert material
        cursor.execute(
            """
            INSERT INTO materials
                (material_id, material_name, category, plant, price, status)
            VALUES
                (%s, %s, %s, %s, %s, %s)
            """,
            (material_id, material_name, category, plant, price, status),
        )

        connection.commit()

        cursor.close()
        connection.close()

        # Return to Materials with previous filter/sort state
        return redirect(
            url_for(
                "materials",
                search=search,
                category=category_filter,
                plant=plant_filter,
                status=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )
        )

    return render_template(
        "material_form.html",
        search=search,
        category_filter=category_filter,
        plant_filter=plant_filter,
        status_filter=status_filter,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
    )


@app.route("/materials/<material_id>/edit", methods=["GET", "POST"])
def edit_material(material_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    category_filter = request.args.get("category", "").strip()
    plant_filter = request.args.get("plant", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "material_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # GET existing material
    # -----------------------------------

    if request.method == "GET":

        cursor.execute(
            """
            SELECT material_id,
                   material_name,
                   category,
                   plant,
                   price,
                   status
            FROM materials
            WHERE material_id = %s
        """,
            (material_id,),
        )

        material = cursor.fetchone()

        cursor.close()
        connection.close()

        if not material:
            return "Material not found", 404

        return render_template(
            "edit_material.html",
            material=material,
            search=search,
            category_filter=category_filter,
            plant_filter=plant_filter,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # POST - Get updated values
    # -----------------------------------

    material_name = request.form.get("material_name", "").strip()
    category = request.form.get("category", "").strip()
    plant = request.form.get("plant", "").strip()
    price = request.form.get("price", "").strip()
    status = request.form.get("status", "").strip()

    # Get preserved state from hidden fields
    search = request.form.get("search", search)
    category_filter = request.form.get("category_filter", category_filter)
    plant_filter = request.form.get("plant_filter", plant_filter)
    status_filter = request.form.get("status_filter", status_filter)
    sort_by = request.form.get("sort_by", sort_by)
    sort_order = request.form.get("sort_order", sort_order)
    page = request.form.get("page", page)
    per_page = request.form.get("per_page", per_page)

    # -----------------------------------
    # Validation
    # -----------------------------------

    if not material_name or not category or not plant or not price or not status:

        cursor.execute(
            """
            SELECT material_id,
                   material_name,
                   category,
                   plant,
                   price,
                   status
            FROM materials
            WHERE material_id = %s
        """,
            (material_id,),
        )

        material = cursor.fetchone()

        cursor.close()
        connection.close()

        return render_template(
            "edit_material.html",
            material=material,
            error="All fields are required.",
            search=search,
            category_filter=category_filter,
            plant_filter=plant_filter,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Price validation
    # -----------------------------------

    try:
        price = float(price)

    except ValueError:

        cursor.execute(
            """
            SELECT material_id,
                   material_name,
                   category,
                   plant,
                   price,
                   status
            FROM materials
            WHERE material_id = %s
        """,
            (material_id,),
        )

        material = cursor.fetchone()

        cursor.close()
        connection.close()

        return render_template(
            "edit_material.html",
            material=material,
            error="Price must be a valid number.",
            search=search,
            category_filter=category_filter,
            plant_filter=plant_filter,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Update database
    # -----------------------------------

    cursor.execute(
        """
        UPDATE materials
        SET material_name = %s,
            category = %s,
            plant = %s,
            price = %s,
            status = %s
        WHERE material_id = %s
    """,
        (material_name, category, plant, price, status, material_id),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "materials",
            search=search,
            category=category_filter,
            plant=plant_filter,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/materials/<material_id>/delete", methods=["POST"])
def delete_material(material_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    plant = request.args.get("plant", "").strip()
    status_filter = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "material_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()

    per_page = request.args.get("per_page", "10").strip()
    page = request.args.get("page", "1").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # Delete material
    cursor.execute(
        """
        DELETE FROM materials
        WHERE material_id = %s
        """,
        (material_id,),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "materials",
            search=search,
            category=category,
            plant=plant,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            per_page=per_page,
            page=page,
        )
    )


@app.route("/materials/export/csv")
def export_materials_csv():

    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    plant = request.args.get("plant", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "material_id")
    sort_order = request.args.get("sort_order", "asc")

    sort_options = {
        "material_id": "material_id",
        "material_name": "CAST(REGEXP_REPLACE(material_name, '[^0-9]', '', 'g') AS INTEGER)",
        "category": "category",
        "plant": "plant",
        "price": "price",
        "status": "status",
    }

    order_column = sort_options.get(sort_by, "material_id")
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT material_id,
               material_name,
               category,
               plant,
               price,
               status
        FROM materials
        WHERE 1=1
    """

    params = []

    if search:
        query += """
            AND (
                material_id ILIKE %s
                OR material_name ILIKE %s
                OR category ILIKE %s
                OR plant ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    if category:
        query += " AND category = %s"
        params.append(category)

    if plant:
        query += " AND plant = %s"
        params.append(plant)

    if status:
        query += " AND status = %s"
        params.append(status)

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    materials = cursor.fetchall()

    cursor.close()
    connection.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        ["Material ID", "Material Name", "Category", "Plant", "Price", "Status"]
    )

    writer.writerows(materials)

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=materials.csv"},
    )


@app.route("/materials/export/excel")
def export_materials_excel():

    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    plant = request.args.get("plant", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "material_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "material_id": "material_id",
        "material_name": "CAST(SUBSTRING(material_name FROM POSITION(' ' IN material_name) + 1) AS INTEGER)",
        "category": "category",
        "plant": "plant",
        "price": "price",
        "status": "status",
    }

    # Prevent invalid column names
    order_column = sort_options.get(sort_by, "material_id")

    # Only allow ASC / DESC
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    # -----------------------------------
    # Database connection
    # -----------------------------------

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT material_id,
               material_name,
               category,
               plant,
               price,
               status
        FROM materials
        WHERE 1=1
    """

    params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:
        query += """
            AND (
                material_id ILIKE %s
                OR material_name ILIKE %s
                OR category ILIKE %s
                OR plant ILIKE %s
                OR status ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # -----------------------------------
    # Category filter
    # -----------------------------------

    if category:
        query += " AND category = %s"
        params.append(category)

    # -----------------------------------
    # Plant filter
    # -----------------------------------

    if plant:
        query += " AND plant = %s"
        params.append(plant)

    # -----------------------------------
    # Status filter
    # -----------------------------------

    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Create Excel workbook
    # -----------------------------------

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Materials"

    # Header
    worksheet.append(
        ["Material ID", "Material Name", "Category", "Plant", "Price", "Status"]
    )

    # Data
    for row in rows:
        worksheet.append(row)

    # -----------------------------------
    # Auto width columns
    # -----------------------------------

    for column in worksheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # -----------------------------------
    # Save workbook to memory
    # -----------------------------------

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    # -----------------------------------
    # Download Excel file
    # -----------------------------------

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="materials.xlsx",
    )


@app.route("/sales_orders")
def sales_orders():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "order_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Pagination
    # -----------------------------------

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    # Allowed page sizes
    allowed_per_page = [10, 25, 50, 100]

    if per_page not in allowed_per_page:
        per_page = 10

    if page < 1:
        page = 1

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "order_id": "order_id",
        "customer_id": "customer_id",
        "material_id": "material_id",
        "quantity": "quantity",
        "status": "status",
        "order_date": "order_date",
    }

    # Prevent invalid column names
    order_column = sort_options.get(sort_by, "order_id")

    # Only allow ASC / DESC
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Total records count
    # -----------------------------------

    count_query = """
        SELECT COUNT(*)
        FROM sales_orders
        WHERE 1=1
    """

    count_params = []

    # Search
    if search:
        count_query += """
            AND (
                order_id ILIKE %s
                OR customer_id ILIKE %s
                OR material_id ILIKE %s
                OR status ILIKE %s
                OR order_date::text ILIKE %s
            )
        """

        count_params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # Status filter
    if status:
        count_query += " AND status = %s"
        count_params.append(status)

    cursor.execute(count_query, count_params)

    total_count = cursor.fetchone()[0]

    # -----------------------------------
    # Calculate total pages
    # -----------------------------------

    total_pages = max(1, (total_count + per_page - 1) // per_page)

    # Prevent invalid page number
    if page > total_pages:
        page = total_pages

    # Calculate offset
    offset = (page - 1) * per_page

    # -----------------------------------
    # Main query
    # -----------------------------------

    query = """
        SELECT order_id,
               customer_id,
               material_id,
               quantity,
               status,
               order_date
        FROM sales_orders
        WHERE 1=1
    """

    params = []

    # Search
    if search:
        query += """
            AND (
                order_id ILIKE %s
                OR customer_id ILIKE %s
                OR material_id ILIKE %s
                OR status ILIKE %s
                OR order_date::text ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # Status filter
    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f"""
        ORDER BY {order_column} {order_direction}
    """

    # -----------------------------------
    # Pagination
    # -----------------------------------

    query += """
        LIMIT %s OFFSET %s
    """

    params.extend([per_page, offset])

    cursor.execute(query, params)

    sales_orders = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Render template
    # -----------------------------------

    return render_template(
        "sales_orders.html",
        sales_orders=sales_orders,
        search=search,
        status=status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
        total_count=total_count,
        total_pages=total_pages,
    )


@app.route("/sales_orders/create", methods=["GET", "POST"])
def create_sales_order():

    # -----------------------------------
    # Preserve filter / sort / pagination
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "order_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Get customers and materials
    # -----------------------------------

    cursor.execute("""
        SELECT customer_id, customer_name
        FROM customers
        ORDER BY customer_id
    """)

    customers = cursor.fetchall()

    cursor.execute("""
        SELECT material_id, material_name
        FROM materials
        ORDER BY material_id
    """)

    materials = cursor.fetchall()

    # -----------------------------------
    # Create Sales Order
    # -----------------------------------

    if request.method == "POST":

        order_id = request.form.get("order_id", "").strip()
        customer_id = request.form.get("customer_id", "").strip()
        material_id = request.form.get("material_id", "").strip()
        quantity = request.form.get("quantity", "").strip()
        new_status = request.form.get("status", "").strip()
        order_date = request.form.get("order_date", "").strip()

        if not all(
            [order_id, customer_id, material_id, quantity, new_status, order_date]
        ):
            cursor.close()
            connection.close()

            return render_template(
                "sales_order_form.html",
                customers=customers,
                materials=materials,
                error="All fields are required.",
                order_id=order_id,
                customer_id=customer_id,
                material_id=material_id,
                quantity=quantity,
                status=new_status,
                order_date=order_date,
                search=search,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )

        try:
            cursor.execute(
                """
                INSERT INTO sales_orders (
                    order_id,
                    customer_id,
                    material_id,
                    quantity,
                    status,
                    order_date
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """,
                (
                    order_id,
                    customer_id,
                    material_id,
                    int(quantity),
                    new_status,
                    order_date,
                ),
            )

            connection.commit()

        except Exception as e:

            connection.rollback()

            cursor.close()
            connection.close()

            return render_template(
                "sales_order_form.html",
                customers=customers,
                materials=materials,
                error=str(e),
                order_id=order_id,
                customer_id=customer_id,
                material_id=material_id,
                quantity=quantity,
                status=new_status,
                order_date=order_date,
                search=search,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )

        cursor.close()
        connection.close()

        # -----------------------------------
        # Return to same filtered/sorted page
        # -----------------------------------

        return redirect(
            url_for(
                "sales_orders",
                search=search,
                status=status_filter,
                sort_by=sort_by,
                sort_order=sort_order,
                page=page,
                per_page=per_page,
            )
        )

    cursor.close()
    connection.close()

    return render_template(
        "sales_order_form.html",
        customers=customers,
        materials=materials,
        search=search,
        status_filter=status_filter,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
    )


@app.route("/sales_orders/<order_id>/edit", methods=["GET", "POST"])
def edit_sales_order(order_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "order_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # GET existing sales order
    # -----------------------------------

    if request.method == "GET":

        cursor.execute(
            """
            SELECT order_id,
                   customer_id,
                   material_id,
                   quantity,
                   status,
                   order_date
            FROM sales_orders
            WHERE order_id = %s
        """,
            (order_id,),
        )

        sales_order = cursor.fetchone()

        # Get existing customers
        cursor.execute("""
            SELECT customer_id
            FROM customers
            ORDER BY customer_id
        """)

        customers = cursor.fetchall()

        # Get existing materials
        cursor.execute("""
            SELECT material_id
            FROM materials
            ORDER BY material_id
        """)

        materials = cursor.fetchall()

        cursor.close()
        connection.close()

        if not sales_order:
            return "Sales Order not found", 404

        return render_template(
            "edit_sales_order.html",
            sales_order=sales_order,
            customers=customers,
            materials=materials,
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # POST - Get updated values
    # -----------------------------------

    customer_id = request.form.get("customer_id", "").strip()
    material_id = request.form.get("material_id", "").strip()
    quantity = request.form.get("quantity", "").strip()
    new_status = request.form.get("status", "").strip()
    order_date = request.form.get("order_date", "").strip()

    # -----------------------------------
    # Get preserved state from hidden fields
    # -----------------------------------

    search = request.form.get("search", search)
    status_filter = request.form.get("status_filter", status_filter)
    sort_by = request.form.get("sort_by", sort_by)
    sort_order = request.form.get("sort_order", sort_order)
    page = request.form.get("page", page)
    per_page = request.form.get("per_page", per_page)

    # -----------------------------------
    # Validation
    # -----------------------------------

    if (
        not customer_id
        or not material_id
        or not quantity
        or not new_status
        or not order_date
    ):

        cursor.execute(
            """
            SELECT order_id,
                   customer_id,
                   material_id,
                   quantity,
                   status,
                   order_date
            FROM sales_orders
            WHERE order_id = %s
        """,
            (order_id,),
        )

        sales_order = cursor.fetchone()

        cursor.execute("""
            SELECT customer_id
            FROM customers
            ORDER BY customer_id
        """)

        customers = cursor.fetchall()

        cursor.execute("""
            SELECT material_id
            FROM materials
            ORDER BY material_id
        """)

        materials = cursor.fetchall()

        cursor.close()
        connection.close()

        return render_template(
            "edit_sales_order.html",
            sales_order=sales_order,
            customers=customers,
            materials=materials,
            error="All fields are required.",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Customer ID validation
    # -----------------------------------

    cursor.execute(
        """
        SELECT customer_id
        FROM customers
        WHERE customer_id = %s
    """,
        (customer_id,),
    )

    customer_exists = cursor.fetchone()

    if not customer_exists:

        cursor.execute(
            """
            SELECT order_id,
                   customer_id,
                   material_id,
                   quantity,
                   status,
                   order_date
            FROM sales_orders
            WHERE order_id = %s
        """,
            (order_id,),
        )

        sales_order = cursor.fetchone()

        cursor.execute("""
            SELECT customer_id
            FROM customers
            ORDER BY customer_id
        """)

        customers = cursor.fetchall()

        cursor.execute("""
            SELECT material_id
            FROM materials
            ORDER BY material_id
        """)

        materials = cursor.fetchall()

        cursor.close()
        connection.close()

        return render_template(
            "edit_sales_order.html",
            sales_order=sales_order,
            customers=customers,
            materials=materials,
            error="Selected Customer ID does not exist.",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Material ID validation
    # -----------------------------------

    cursor.execute(
        """
        SELECT material_id
        FROM materials
        WHERE material_id = %s
    """,
        (material_id,),
    )

    material_exists = cursor.fetchone()

    if not material_exists:

        cursor.execute(
            """
            SELECT order_id,
                   customer_id,
                   material_id,
                   quantity,
                   status,
                   order_date
            FROM sales_orders
            WHERE order_id = %s
        """,
            (order_id,),
        )

        sales_order = cursor.fetchone()

        cursor.execute("""
            SELECT customer_id
            FROM customers
            ORDER BY customer_id
        """)

        customers = cursor.fetchall()

        cursor.execute("""
            SELECT material_id
            FROM materials
            ORDER BY material_id
        """)

        materials = cursor.fetchall()

        cursor.close()
        connection.close()

        return render_template(
            "edit_sales_order.html",
            sales_order=sales_order,
            customers=customers,
            materials=materials,
            error="Selected Material ID does not exist.",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Quantity validation
    # -----------------------------------

    try:
        quantity = int(quantity)

        if quantity <= 0:
            raise ValueError

    except ValueError:

        cursor.execute(
            """
            SELECT order_id,
                   customer_id,
                   material_id,
                   quantity,
                   status,
                   order_date
            FROM sales_orders
            WHERE order_id = %s
        """,
            (order_id,),
        )

        sales_order = cursor.fetchone()

        cursor.execute("""
            SELECT customer_id
            FROM customers
            ORDER BY customer_id
        """)

        customers = cursor.fetchall()

        cursor.execute("""
            SELECT material_id
            FROM materials
            ORDER BY material_id
        """)

        materials = cursor.fetchall()

        cursor.close()
        connection.close()

        return render_template(
            "edit_sales_order.html",
            sales_order=sales_order,
            customers=customers,
            materials=materials,
            error="Quantity must be a positive number.",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Update database
    # -----------------------------------

    cursor.execute(
        """
        UPDATE sales_orders
        SET customer_id = %s,
            material_id = %s,
            quantity = %s,
            status = %s,
            order_date = %s
        WHERE order_id = %s
    """,
        (customer_id, material_id, quantity, new_status, order_date, order_id),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "sales_orders",
            search=search,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/sales_orders/<order_id>/delete", methods=["POST"])
def delete_sales_order(order_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "order_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Delete sales order
    # -----------------------------------

    cursor.execute(
        """
        DELETE FROM sales_orders
        WHERE order_id = %s
        """,
        (order_id,),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "sales_orders",
            search=search,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/sales_orders/export/csv")
def export_sales_orders_csv():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "order_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "order_id": "order_id",
        "customer_id": "customer_id",
        "material_id": "material_id",
        "quantity": "quantity",
        "status": "status",
        "order_date": "order_date",
    }

    order_column = sort_options.get(sort_by, "order_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    # -----------------------------------
    # Database connection
    # -----------------------------------

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT order_id,
               customer_id,
               material_id,
               quantity,
               status,
               order_date
        FROM sales_orders
        WHERE 1=1
    """

    params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:
        query += """
            AND (
                order_id ILIKE %s
                OR customer_id ILIKE %s
                OR material_id ILIKE %s
                OR status ILIKE %s
                OR order_date::text ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # -----------------------------------
    # Status filter
    # -----------------------------------

    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Create CSV
    # -----------------------------------

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow(
        [
            "Sales Order ID",
            "Customer ID",
            "Material ID",
            "Quantity",
            "Status",
            "Order Date",
        ]
    )

    for row in rows:
        writer.writerow(row)

    # -----------------------------------
    # Return CSV
    # -----------------------------------

    response = Response(output.getvalue(), mimetype="text/csv")

    response.headers["Content-Disposition"] = "attachment; filename=sales_orders.csv"

    return response


@app.route("/sales_orders/export/excel")
def export_sales_orders_excel():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "order_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "order_id": "order_id",
        "customer_id": "customer_id",
        "material_id": "material_id",
        "quantity": "quantity",
        "status": "status",
        "order_date": "order_date",
    }

    order_column = sort_options.get(sort_by, "order_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    # -----------------------------------
    # Database connection
    # -----------------------------------

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT order_id,
               customer_id,
               material_id,
               quantity,
               status,
               order_date
        FROM sales_orders
        WHERE 1=1
    """

    params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:
        query += """
            AND (
                order_id ILIKE %s
                OR customer_id ILIKE %s
                OR material_id ILIKE %s
                OR status ILIKE %s
                OR order_date::text ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # -----------------------------------
    # Status filter
    # -----------------------------------

    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Create Excel workbook
    # -----------------------------------

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Sales Orders"

    # Header
    worksheet.append(
        [
            "Sales Order ID",
            "Customer ID",
            "Material ID",
            "Quantity",
            "Status",
            "Order Date",
        ]
    )

    # Data
    for row in rows:
        worksheet.append(row)

    # -----------------------------------
    # Auto width columns
    # -----------------------------------

    for column in worksheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # -----------------------------------
    # Save Excel to memory
    # -----------------------------------

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    # -----------------------------------
    # Download Excel
    # -----------------------------------

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="sales_orders.xlsx",
    )


@app.route("/support_requests")
def support_requests():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "request_id")
    sort_order = request.args.get("sort_order", "asc")

    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    # Allowed rows per page
    allowed_per_page = [10, 25, 50, 100]

    if per_page not in allowed_per_page:
        per_page = 10

    if page < 1:
        page = 1

    # Allowed columns for sorting
    sort_options = {
        "request_id": "request_id",
        "issue_type": "issue_type",
        "description": "description",
        "status": "status",
        "created_date": "created_date",
    }

    # Prevent invalid column names
    order_column = sort_options.get(sort_by, "request_id")

    # Only allow ASC / DESC
    order_direction = "DESC" if sort_order == "desc" else "ASC"

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Main Query
    # -----------------------------------

    query = """
        SELECT request_id,
               issue_type,
               description,
               status,
               created_date
        FROM support_requests
        WHERE 1=1
    """

    # -----------------------------------
    # Count Query
    # -----------------------------------

    count_query = """
        SELECT COUNT(*)
        FROM support_requests
        WHERE 1=1
    """

    params = []
    count_params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:

        search_condition = """
            AND (
                request_id ILIKE %s
                OR issue_type ILIKE %s
                OR description ILIKE %s
                OR status ILIKE %s
                OR created_date::text ILIKE %s
            )
        """

        query += search_condition
        count_query += search_condition

        search_params = [
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
        ]

        params.extend(search_params)
        count_params.extend(search_params)

    # -----------------------------------
    # Status Filter
    # -----------------------------------

    if status:

        query += " AND status = %s"
        count_query += " AND status = %s"

        params.append(status)
        count_params.append(status)

    # -----------------------------------
    # Total Records
    # -----------------------------------

    cursor.execute(count_query, count_params)

    total_records = cursor.fetchone()[0]

    # -----------------------------------
    # Total Pages
    # -----------------------------------

    total_pages = max(1, (total_records + per_page - 1) // per_page)

    # Prevent invalid page number
    if page > total_pages:
        page = total_pages

    # -----------------------------------
    # Pagination Offset
    # -----------------------------------

    offset = (page - 1) * per_page

    # -----------------------------------
    # Sorting + Pagination
    # -----------------------------------

    query += f"""
        ORDER BY {order_column} {order_direction}
        LIMIT %s OFFSET %s
    """

    params.extend([per_page, offset])

    cursor.execute(query, params)

    support_requests = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "support_requests.html",
        support_requests=support_requests,
        search=search,
        status=status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
        total_records=total_records,
        total_pages=total_pages,
    )


@app.route("/support_requests/create", methods=["GET", "POST"])
def create_support_request():

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "request_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # GET - Show Add Form
    # -----------------------------------

    if request.method == "GET":

        cursor.close()
        connection.close()

        return render_template(
            "support_request_form.html",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # POST - Get form values
    # -----------------------------------

    request_id = request.form.get("request_id", "").strip()
    issue_type = request.form.get("issue_type", "").strip()
    description = request.form.get("description", "").strip()
    status = request.form.get("status", "").strip()
    created_date = request.form.get("created_date", "").strip()

    # -----------------------------------
    # Get preserved state from hidden fields
    # -----------------------------------

    search = request.form.get("search", search)
    status_filter = request.form.get("status_filter", status_filter)
    sort_by = request.form.get("sort_by", sort_by)
    sort_order = request.form.get("sort_order", sort_order)
    page = request.form.get("page", page)
    per_page = request.form.get("per_page", per_page)

    # -----------------------------------
    # Validation
    # -----------------------------------

    if (
        not request_id
        or not issue_type
        or not description
        or not status
        or not created_date
    ):

        cursor.close()
        connection.close()

        return render_template(
            "support_request_form.html",
            error="All fields are required.",
            request_id=request_id,
            issue_type=issue_type,
            description=description,
            status=status,
            created_date=created_date,
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Check duplicate Request ID
    # -----------------------------------

    cursor.execute(
        """
        SELECT 1
        FROM support_requests
        WHERE request_id = %s
        """,
        (request_id,),
    )

    existing_request = cursor.fetchone()

    if existing_request:

        cursor.close()
        connection.close()

        return render_template(
            "support_request_form.html",
            error="Request ID already exists. Please enter a different Request ID.",
            request_id=request_id,
            issue_type=issue_type,
            description=description,
            status=status,
            created_date=created_date,
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Insert into database
    # -----------------------------------

    cursor.execute(
        """
        INSERT INTO support_requests (
            request_id,
            issue_type,
            description,
            status,
            created_date
        )
        VALUES (%s, %s, %s, %s, %s)
        """,
        (
            request_id,
            issue_type,
            description,
            status,
            created_date,
        ),
    )

    connection.commit()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "support_requests",
            search=search,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/support_requests/<request_id>/edit", methods=["GET", "POST"])
def edit_support_request(request_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "request_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # GET - Show Edit Form
    # -----------------------------------

    if request.method == "GET":

        cursor.execute(
            """
            SELECT request_id,
                   issue_type,
                   description,
                   status,
                   created_date
            FROM support_requests
            WHERE request_id = %s
            """,
            (request_id,),
        )

        support_request = cursor.fetchone()

        cursor.close()
        connection.close()

        if not support_request:
            return "Support Request not found", 404

        return render_template(
            "edit_support_request.html",
            support_request=support_request,
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # POST - Get updated values
    # -----------------------------------

    issue_type = request.form.get("issue_type", "").strip()
    description = request.form.get("description", "").strip()
    new_status = request.form.get("status", "").strip()
    created_date = request.form.get("created_date", "").strip()

    # -----------------------------------
    # Get preserved state from hidden fields
    # -----------------------------------

    search = request.form.get("search", search)
    status_filter = request.form.get("status_filter", status_filter)
    sort_by = request.form.get("sort_by", sort_by)
    sort_order = request.form.get("sort_order", sort_order)
    page = request.form.get("page", page)
    per_page = request.form.get("per_page", per_page)

    # -----------------------------------
    # Validation
    # -----------------------------------

    if not issue_type or not description or not new_status or not created_date:

        cursor.execute(
            """
            SELECT request_id,
                   issue_type,
                   description,
                   status,
                   created_date
            FROM support_requests
            WHERE request_id = %s
            """,
            (request_id,),
        )

        support_request = cursor.fetchone()

        cursor.close()
        connection.close()

        return render_template(
            "edit_support_request.html",
            support_request=support_request,
            error="All fields are required.",
            search=search,
            status_filter=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )

    # -----------------------------------
    # Update database
    # -----------------------------------

    cursor.execute(
        """
        UPDATE support_requests
        SET issue_type = %s,
            description = %s,
            status = %s,
            created_date = %s
        WHERE request_id = %s
        """,
        (
            issue_type,
            description,
            new_status,
            created_date,
            request_id,
        ),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "support_requests",
            search=search,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/support_requests/<request_id>/delete", methods=["POST"])
def delete_support_request(request_id):

    # -----------------------------------
    # Preserve filter / sort parameters
    # -----------------------------------

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    sort_by = request.args.get("sort_by", "request_id").strip()
    sort_order = request.args.get("sort_order", "asc").strip()
    page = request.args.get("page", "1").strip()
    per_page = request.args.get("per_page", "10").strip()

    connection = get_connection()
    cursor = connection.cursor()

    # -----------------------------------
    # Delete support request
    # -----------------------------------

    cursor.execute(
        """
        DELETE FROM support_requests
        WHERE request_id = %s
        """,
        (request_id,),
    )

    connection.commit()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Return to same filtered/sorted page
    # -----------------------------------

    return redirect(
        url_for(
            "support_requests",
            search=search,
            status=status_filter,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
        )
    )


@app.route("/support_requests/export/csv")
def export_support_requests_csv():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "request_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "request_id": "request_id",
        "issue_type": "issue_type",
        "description": "description",
        "status": "status",
        "created_date": "created_date",
    }

    order_column = sort_options.get(sort_by, "request_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    # -----------------------------------
    # Database connection
    # -----------------------------------

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT request_id,
               issue_type,
               description,
               status,
               created_date
        FROM support_requests
        WHERE 1=1
    """

    params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:
        query += """
            AND (
                request_id ILIKE %s
                OR issue_type ILIKE %s
                OR description ILIKE %s
                OR status ILIKE %s
                OR created_date::text ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # -----------------------------------
    # Status filter
    # -----------------------------------

    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Create CSV
    # -----------------------------------

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow(
        ["Request ID", "Issue Type", "Description", "Status", "Created Date"]
    )

    for row in rows:
        writer.writerow(row)

    # -----------------------------------
    # Return CSV
    # -----------------------------------

    response = Response(output.getvalue(), mimetype="text/csv")

    response.headers["Content-Disposition"] = (
        "attachment; filename=support_requests.csv"
    )

    return response


@app.route("/support_requests/export/excel")
def export_support_requests_excel():

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()

    sort_by = request.args.get("sort_by", "request_id")
    sort_order = request.args.get("sort_order", "asc")

    # -----------------------------------
    # Allowed columns for sorting
    # -----------------------------------

    sort_options = {
        "request_id": "request_id",
        "issue_type": "issue_type",
        "description": "description",
        "status": "status",
        "created_date": "created_date",
    }

    order_column = sort_options.get(sort_by, "request_id")

    order_direction = "DESC" if sort_order == "desc" else "ASC"

    # -----------------------------------
    # Database connection
    # -----------------------------------

    connection = get_connection()
    cursor = connection.cursor()

    query = """
        SELECT request_id,
               issue_type,
               description,
               status,
               created_date
        FROM support_requests
        WHERE 1=1
    """

    params = []

    # -----------------------------------
    # Search
    # -----------------------------------

    if search:
        query += """
            AND (
                request_id ILIKE %s
                OR issue_type ILIKE %s
                OR description ILIKE %s
                OR status ILIKE %s
                OR created_date::text ILIKE %s
            )
        """

        params.extend(
            [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
        )

    # -----------------------------------
    # Status filter
    # -----------------------------------

    if status:
        query += " AND status = %s"
        params.append(status)

    # -----------------------------------
    # Sorting
    # -----------------------------------

    query += f" ORDER BY {order_column} {order_direction}"

    cursor.execute(query, params)

    rows = cursor.fetchall()

    cursor.close()
    connection.close()

    # -----------------------------------
    # Create Excel workbook
    # -----------------------------------

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Support Requests"

    # Header
    worksheet.append(
        ["Request ID", "Issue Type", "Description", "Status", "Created Date"]
    )

    # Data
    for row in rows:
        worksheet.append(row)

    # -----------------------------------
    # Auto width columns
    # -----------------------------------

    for column in worksheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 2

    # -----------------------------------
    # Save Excel to memory
    # -----------------------------------

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    # -----------------------------------
    # Download Excel
    # -----------------------------------

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="support_requests.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True)
